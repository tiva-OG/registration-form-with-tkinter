from tkinter import (Entry, Frame, Label, Radiobutton, StringVar, Tk,
                     messagebox, ttk)

import openpyxl

from validations import days_of_month, no_name


def callbackYear(event):
    year = year_combo.get()
    month = month_combo.get()
    day = day_combo.get()
    
    days = days_of_month(year, month)
    day_choice = no_name(year, month, day)
    day_combo.set(day_choice)
    day_combo.configure(values=days)

def callbackMonth(event):
    year = year_combo.get()
    month = month_combo.get()
    day = day_combo.get()
    
    days = days_of_month(year, month)
    day_choice = no_name(year, month, day)
    day_combo.set(day_choice)
    day_combo.configure(values=days)

class Database():
    
    def __init__(self, data, wb_path=".\\test.xlsx"):
        
        self.wb_path = wb_path
        self.workbook = openpyxl.load_workbook(wb_path)
        self.sheet = self.workbook.active
        
        # Column names
        self.sheet.cell(row=1, column=1).value = "Name"
        self.sheet.cell(row=1, column=2).value = "State"
        self.sheet.cell(row=1, column=3).value = "Phone"
        self.sheet.cell(row=1, column=4).value = "Email"
        self.sheet.cell(row=1, column=5).value = "Gender"
        self.sheet.cell(row=1, column=6).value = "DOB"
        self.sheet.cell(row=1, column=7).value = "Interest"
        
        # initialize data collection
        self.name = data["name"]
        self.state = data["state"]
        self.phone = data["phone"]
        self.email = data["email"]
        self.gender = data["gender"]
        self.interest = data["interest"]
        
        self.day = data["dob"]["day"]
        self.month = data["dob"]["month"]
        self.year = data["dob"]["year"]
        self.dob = self.day.get() + "/" + self.month.get() + "/" + self.year.get()
    
    def insert(self):
        
        # insert data from GUI window to spreadsheet
        if any(v.get() == v.placeholder for v in [self.name, self.phone, self.email]):
            return "Incomplete Field(s)"
        
        else:
            curr_row = self.sheet.max_row + 1
            
            self.sheet.cell(row=curr_row, column=1).value = self.name.get()
            self.sheet.cell(row=curr_row, column=2).value = self.state.get()
            self.sheet.cell(row=curr_row, column=3).value = self.phone.get()
            self.sheet.cell(row=curr_row, column=4).value = self.email.get()
            self.sheet.cell(row=curr_row, column=5).value = self.gender.get()
            self.sheet.cell(row=curr_row, column=6).value = self.dob
            self.sheet.cell(row=curr_row, column=7).value = self.interest.get()
        
            self.workbook.save(self.wb_path)
            return "Registration Complete"
    
    def reset(self):
        self.name.delete(0, "end")
        self.name.put_placeholder()
        self.state.set("Abia")
        self.phone.delete(0, "end")
        self.phone.put_placeholder()
        self.email.delete(0, "end")
        self.email.put_placeholder()
        self.gender.set("Male")
        self.year.set(1980)
        self.month.set("January")
        self.day.set(1)
        self.interest.set("Data Science")

class EntryWithPlaceholder(Entry):
    
    def __init__(self, master=None, placeholder="PLACEHOLDER", color="grey", **kwargs):
        super().__init__(master, **kwargs)
        
        self.placeholder = placeholder
        self.placeholder_color = color
        self.default_fg_color = self["fg"]
        
        self.bind("<FocusIn>", self.focus_in)
        self.bind("<FocusOut>", self.focus_out)
        
        self.put_placeholder()
    
    def put_placeholder(self):
        self.insert(0, self.placeholder)
        self["fg"] = self.placeholder_color
    
    def focus_in(self, *args):
        if self["fg"] == self.placeholder_color:
            self.delete(0, "end")
            self["fg"] = self.default_fg_color
    
    def focus_out(self, *args):
        if not self.get().strip():
            self.put_placeholder()

class RegistrationForm(Tk):
    
    def __init__(self):
        super().__init__()
        self.initializeUI()
    
    def initializeUI(self):
        self.title("BB-HUB Registration Form")
        self.minsize(300, 200)
        self.geometry("400x650+50+100")
        self.configure(background="lightgrey")
        self.setup_window()
    
    def register(self):
        database = Database(data=self.data)
        feedback = database.insert()
        
        if feedback == "Incomplete Field(s)":
            messagebox.showwarning("BB-HUB", message=feedback)
        else:
            messagebox.showinfo("BB-HUB", message=feedback)
            database.reset()
    
    def setup_window(self):
        '''Set up widgets'''
        
        global year_combo, month_combo, day_combo, gender_choice
        bg_color = "lightgrey"
        
        years = list(range(1980, 2007))
        days = list(range(1, 32))
        months = [
            "January", "February", "March", "April", 
            "May", "June", "July", "August","September", 
            "October", "November", "December"
            ]
        states = [
            "Abia", "Adamawa", "Akwa-Ibom", "Anambra", 
            "Bauchi", "Bayelsa", "Benue", "Borno", 
            "Cross-River", "Delta", "Ebonyi", "Edo", 
            "Ekiti", "Enugu", "Gombe", "Imo", 
            "Jigawa", "Kaduna", "Kano", "Katsina", 
            "Kebbi", "Kogi", "Kwara", "Lagos", 
            "Nasarawa", "Niger", "Ogun", "Ondo", 
            "Osun", "Oyo", "Plateau", "Rivers", 
            "Sokoto", "Taraba", "Yobe", "Zamfara"
            ]
        interests = [
            "Data Science", "Web Development", 
            "Machine Learning", "Python", 
            "DB Management", "Mobile Development"
            ]
        
        
        heading = Label(self, 
                        text="Welcome to the BB-HUB\nPlease enter your information.", 
                        bg="lightgrey", 
                        font=('Courier New', 9),
                        bd=10)
        heading.pack(pady=20)
        
        line = ttk.Separator(self, orient="horizontal")
        line.pack(fill='x')
        
        name = Label(self, text="Name:", bg="lightgrey")
        name.pack(anchor="w", padx=10, pady=(10, 0))
        name_field = EntryWithPlaceholder(self, placeholder=" surname first", width=50)
        name_field.pack(anchor="w", padx=10, pady=(5, 10))
        
        state = Label(self, text="State-of-Origin:", bg="lightgrey")
        state.pack(anchor="w", padx=10, pady=(10, 0))
        state_combo = ttk.Combobox(self, state="readonly", values=states)
        state_combo.set("Abia")
        state_combo.pack(anchor="w", padx=10, pady=(5, 10))
        
        phone = Label(self, text="Phone Number:", bg="lightgrey")
        phone.pack(anchor="w", padx=10, pady=(10, 0))
        phone_field =  EntryWithPlaceholder(self, placeholder=" ex: 080******678", width=50)
        phone_field.pack(anchor="w", padx=10, pady=(5, 10))
        
        email = Label(self, text="Email Address:", bg="lightgrey")
        email.pack(anchor="w", padx=10, pady=(10, 0))
        email_field =  EntryWithPlaceholder(self, placeholder=" ex: myname@email.com", width=50)
        email_field.pack(anchor="w", padx=10, pady=(5, 10))
        
        gender = Label(self, text="Gender:", bg="lightgrey")
        gender.pack(anchor="w", padx=10, pady=(10, 0))
        gender_frame = Frame(self, bg=bg_color)
        gender_frame.pack(anchor="w", padx=10, pady=(5, 10))
        gender_choice = StringVar(self, "Male")
        Radiobutton(gender_frame, text="Male", value="Male", variable=gender_choice, bg=bg_color).pack(side="left")
        Radiobutton(gender_frame, text="Female", value="Female" ,variable=gender_choice, bg=bg_color).pack(side="left", padx=15)
        
        dob = Label(self, text="Date-of-Birth:", bg=bg_color)
        dob.pack(anchor="w", padx=10, pady=(10, 0))
        dob_frame = Frame(self, bg=bg_color)
        dob_frame.pack(anchor="w", padx=10, pady=(5, 10))
        
        year_combo = ttk.Combobox(dob_frame, state="readonly", values=years, width=5)
        year_combo.set(1980)
        year_combo.pack(side="left", padx=(0, 10))
        month_combo = ttk.Combobox(dob_frame, state="readonly", values=months, width=11)
        month_combo.set("January")
        month_combo.pack(side="left", padx=(0, 10))
        day_combo = ttk.Combobox(dob_frame, state="readonly", values=days, width=4)
        day_combo.set(1)
        day_combo.pack(side="left", padx=(0, 10))
        
        year_combo.bind("<<ComboboxSelected>>", callbackYear)
        month_combo.bind("<<ComboboxSelected>>", callbackMonth)
        
        interest = Label(self, text="Interests: ", bg=bg_color)
        interest.pack(anchor="w", padx=10, pady=(10, 0))
        interest_combo = ttk.Combobox(state="readonly", values=interests)
        interest_combo.set("Data Science")
        interest_combo.pack(anchor="w", padx=10, pady=(5, 10))
        
        self.data = {
            "name": name_field,
            "state": state_combo,
            "phone": phone_field,
            "email": email_field,
            "gender": gender_choice,
            "dob": {
                "day": day_combo,
                "month": month_combo,
                "year": year_combo
                },
            "interest": interest_combo
        }
        
        style = ttk.Style()
        style.configure("custom.TButton", bg="skyblue", fg="white", font="Courier 10 underline")
        
        register = ttk.Button(self, text="Register", style="custom.TButton", command=self.register)
        register.pack(pady=20)

if __name__ == "__main__":
    
    app = RegistrationForm()
    app.mainloop()
    